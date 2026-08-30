import pandas as pd
import io
from datetime import date
from typing import List, Dict, Any, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import StreamingResponse
from app.utils.jalali import parse_jalali_date

class ExcelService:
    @staticmethod
    def generate_template(import_type: str) -> bytes:
        """تولید قالب Excel بر اساس نوع عملیات"""
        wb = Workbook()
        ws = wb.active
        
        # ستون‌ها بر اساس نوع
        if import_type == "DEBIT":
            columns = ["شماره عضو", "نام کامل", "مبلغ (ریال)", "شرح"]
            sample_data = [
                ["010001", "علی محمدی", "50000000", "قسط اول"],
                ["010002", "رضا احمدی", "30000000", "قسط اول"],
            ]
        elif import_type == "CREDIT":
            columns = ["شماره عضو", "نام کامل", "مبلغ (ریال)", "شرح"]
            sample_data = [
                ["010001", "علی محمدی", "30000000", "وام خرید"],
                ["010002", "رضا احمدی", "25000000", "وام خرید"],
            ]
        elif import_type == "MEMBER":
            columns = ["شماره عضو", "نام کامل", "کد ملی", "موبایل", "تلفن", "آدرس"]
            sample_data = [
                ["010001", "علی محمدی", "1234567890", "09121234567", "", ""],
                ["010002", "رضا احمدی", "9876543210", "09127654321", "", ""],
            ]
        elif import_type == "BANK_STATEMENT":
            columns = ["تاریخ", "شماره حساب", "مبلغ (ریال)", "نوع", "شماره مرجع", "شرح"]
            sample_data = [
                ["1404-05-03", "1234567890", "50000000", "DEPOSIT", "REF-001", "واریز نقدی"],
                ["1404-05-03", "1234567890", "30000000", "WITHDRAWAL", "REF-002", "برداشت"],
            ]
        else:
            columns = ["شماره عضو", "مبلغ (ریال)", "شرح"]
            sample_data = [["010001", "50000000", "توضیحات"]]
        
        # نوشتن هدر
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # نوشتن نمونه داده
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # تنظیم عرض ستون‌ها
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # ذخیره در حافظه
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def parse_excel(file_content: bytes, date_calendar: str = "jalali") -> Tuple[List[Dict[str, Any]], List[str]]:
        """خواندن فایل Excel و تبدیل به لیست دیکشنری"""
        if date_calendar not in {"jalali", "gregorian"}:
            return [], ["تقویم تاریخ باید jalali یا gregorian باشد"]

        try:
            df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
        except Exception as e:
            return [], [f"خطا در خواندن فایل Excel: {str(e)}"]

        data = df.to_dict(orient='records')

        # استانداردسازی نام ستون‌ها
        column_map = {
            "شماره عضو": "member_no",
            "نام کامل": "full_name",
            "مبلغ (ریال)": "amount",
            "شرح": "description",
            "تاریخ": "date",
            "شماره حساب": "account_no",
            "نوع": "transaction_type",
            "شماره مرجع": "reference_no",
            "کد ملی": "national_code",
            "موبایل": "mobile",
            "تلفن": "phone",
            "آدرس": "address"
        }

        result = []
        errors = []
        for row_number, row in enumerate(data, start=2):
            item = {}
            row_errors = []

            # استانداردسازی نام ستون‌ها
            for persian_col, english_col in column_map.items():
                if persian_col not in row or pd.isna(row[persian_col]):
                    continue

                value = row[persian_col]
                if english_col == "amount":
                    try:
                        value = int(float(str(value).replace(",", "")))
                    except (TypeError, ValueError):
                        row_errors.append(f"ردیف {row_number}: مبلغ نامعتبر است")
                        value = None
                elif english_col == "date":
                    if isinstance(value, pd.Timestamp):
                        if date_calendar == "gregorian":
                            value = value.date()
                        else:
                            row_errors.append(
                                f"ردیف {row_number}: تاریخ Excel-native فقط در حالت میلادی پشتیبانی می‌شود"
                            )
                            value = None
                    elif isinstance(value, date):
                        value = value if date_calendar == "gregorian" else None
                        if value is None:
                            row_errors.append(
                                f"ردیف {row_number}: تاریخ date فقط در حالت میلادی پشتیبانی می‌شود"
                            )
                    elif isinstance(value, str):
                        normalized_value = value.strip()
                        try:
                            if date_calendar == "gregorian":
                                value = date.fromisoformat(normalized_value)
                            else:
                                value = parse_jalali_date(normalized_value)
                        except ValueError:
                            if date_calendar == "gregorian":
                                row_errors.append(
                                    f"ردیف {row_number}: تاریخ میلادی باید به شکل YYYY-MM-DD باشد"
                                )
                            else:
                                row_errors.append(f"ردیف {row_number}: تاریخ جلالی معتبر نیست")
                            value = None
                    else:
                        row_errors.append(f"ردیف {row_number}: نوع داده تاریخ پشتیبانی نمی‌شود")
                        value = None
                elif english_col in {"member_no", "account_no", "transaction_type", "reference_no"}:
                    value = str(value).strip()

                item[english_col] = value

            result.append(item)
            errors.extend(row_errors)

        return result, errors
    
    @staticmethod
    def create_excel_response(data: List[Dict[str, Any]], columns: List[str], filename: str):
        """ایجاد پاسخ Excel برای دانلود"""
        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
